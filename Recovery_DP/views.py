from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import *
from modelmasterapp.models import *
from django.contrib import messages
from datetime import datetime
from django.utils.safestring import mark_safe
import re
import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.parsers import MultiPartParser
from rest_framework import status
import openpyxl
import re
from datetime import datetime
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.db.models import OuterRef, Subquery
import json
from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.parsers import MultiPartParser, JSONParser
import math
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.db import transaction, IntegrityError
from django.utils.timezone import now
from django.http import JsonResponse


class REC_DPBulkUploadView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Recovery_DP/Recovery_DP_BulkUpload.html'
    parser_classes = [MultiPartParser, JSONParser]

    def validate_codes(self, stock_raw, model_field):
        model_match = re.match(r'^(\d+)[A-Za-z]', stock_raw)
        if not model_match:
            return None, None, "❌ Cannot extract model number from: {}".format(stock_raw)

        model_no = model_match.group(1)
        model_stock = ModelMaster.objects.filter(model_no=model_no).first()
        if not model_stock:
            return None, None, f"❌ Model number '{model_no}' not found in ModelMaster."

        print(f"✅ Model number '{model_no}' exists in ModelMaster.")

        letters = ''.join(re.findall(r'[A-Za-z]', model_field))
        if len(letters) < 3:
            return None, None, f"❌ Invalid model format. Found less than 3 letters in: {model_field}"

        return model_stock, letters[:3], None

    def get(self, request, format=None):
        master_data = RecoveryMasterCreation.objects.all()
        return Response({'master_data': master_data})

    def post(self, request, format=None):
        # Check if request contains JSON data (from datatable)
        if request.content_type == 'application/json':
            return self.handle_datatable_submission(request)
        
        # Handle file upload (existing functionality)
        return self.handle_file_upload(request)

    def handle_datatable_submission(self, request):
        """Handle submission from HTML datatable"""
        try:
            # Parse JSON data from request body
            data = request.data
            rows = data.get('rows', [])
            
            if not rows:
                return JsonResponse({
                    'success': False,
                    'error': '❌ No data provided for processing.'
                }, status=400)

            success_count = 0
            failure_count = 0
            failed_rows = []

            for idx, row_data in enumerate(rows, start=1):
                try:
                    # Extract data from row
                    stock_raw = str(row_data.get('Stock No', '')).strip()
                    model_field = str(row_data.get('Model', '')).strip()
                    plating_color_text = str(row_data.get('Plating Color', '')).strip()
                    qty = row_data.get('Quantity')
                    s_loc = str(row_data.get('S Loc', '')).strip()

                    print(f"\n🔄 Processing row {idx}: Stock={stock_raw}, Model={model_field}, Color={plating_color_text}, Qty={qty}, Loc={s_loc}")

                    # Validate required fields
                    empty_fields = []
                    if not stock_raw:
                        empty_fields.append("Stock No")
                    if not model_field:
                        empty_fields.append("Model Code")
                    if not plating_color_text:
                        empty_fields.append("Plating Color")
                    if qty in [None, '', 0]:
                        empty_fields.append("Total Quantity")
                    if not s_loc:
                        empty_fields.append("Location")

                    if empty_fields:
                        field_list = ", ".join(empty_fields)
                        failed_rows.append(f"Row {idx}: ❌ {field_list} should not be empty.")
                        failure_count += 1
                        continue

                    # Convert quantity to integer
                    try:
                        qty = int(qty)
                    except (ValueError, TypeError):
                        failed_rows.append(f"Row {idx}: ❌ Invalid quantity value: {qty}")
                        failure_count += 1
                        continue

                    # Validate codes
                    model_stock, codes, error_msg = self.validate_codes(stock_raw, model_field)
                    if error_msg:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: {error_msg}")
                        continue

                    if not codes:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Could not extract codes from model field.")
                        continue

                    plating_code, polish_code, version_code = codes
                    print(f"🧪 Model Code Extracted ➤ Plating: '{plating_code}', Polish: '{polish_code}', Version: '{version_code}'")

                    # Validate plating code
                    plating_obj_code = Plating_Color.objects.filter(plating_color_internal=plating_code).first()
                    if not plating_obj_code:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Plating code '{plating_code}' not found.")
                        continue

                    # Validate plating color from Excel
                    plating_color_obj_excel = Plating_Color.objects.filter(plating_color=plating_color_text).first()
                    if not plating_color_obj_excel:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Plating Color '{plating_color_text}' not found.")
                        continue

                    # Validate polish code
                    polish_obj = PolishFinishType.objects.filter(polish_internal=polish_code).first()
                    if not polish_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Polish code '{polish_code}' not found.")
                        continue

                    # Validate version code
                    version_obj = Version.objects.filter(version_internal=version_code).first()
                    if not version_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Version code '{version_code}' not found.")
                        continue

                    # Validate location format
                    if "_" not in s_loc:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Invalid s_loc format: '{s_loc}'")
                        continue

                    vendor_name, loc_name = s_loc.split("_", 1)
                    
                    # Validate vendor
                    vendor_obj = Vendor.objects.filter(vendor_name=vendor_name).first()
                    if not vendor_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Vendor '{vendor_name}' not found.")
                        continue

                    # Validate location
                    location_obj = Location.objects.filter(location_name=loc_name).first()
                    if not location_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Location '{loc_name}' not found.")
                        continue

                    # Generate batch ID
                    batch_id = f"BATCH-{datetime.now().strftime('%Y%m%d%H%M%S')}-{idx}"

                    # Create RecoveryMasterCreation record
                    RecoveryMasterCreation.objects.create(
                        batch_id=batch_id,
                        model_stock_no=model_stock,
                        plating_color=plating_color_obj_excel.plating_color,
                        vendor_internal=vendor_obj.vendor_internal,
                        location=location_obj,
                        tray_capacity=model_stock.tray_capacity if model_stock else None,
                        tray_type=model_stock.tray_type.tray_type if model_stock and model_stock.tray_type else None,
                        gender=model_stock.gender if model_stock else None,
                        ep_bath_type=model_stock.ep_bath_type if model_stock else None,
                        brand=model_stock.brand if model_stock else None,
                        total_batch_quantity=qty,
                        version=version_obj if version_obj else None,
                        polish_finish=polish_obj if polish_obj else None,
                    )
                    print("✅ Row saved successfully!")
                    success_count += 1

                except Exception as e:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Error processing row: {str(e).splitlines()[0]}")
                    print(f"❌ Error processing row {idx}: {str(e)}")

            # Prepare response
            if success_count > 0 and failure_count == 0:
                return JsonResponse({
                    'success': True,
                    'message': f"✅ {success_count} row(s) processed successfully."
                })
            elif success_count > 0 and failure_count > 0:
                return JsonResponse({
                    'success': True,
                    'message': f"⚠️ Partial Success: {success_count} succeeded, {failure_count} failed.",
                    'failed_rows': failed_rows
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': f"❌ All {failure_count} row(s) failed to process.",
                    'failed_rows': failed_rows
                }, status=400)

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': '❌ Invalid JSON data provided.'
            }, status=400)
        except Exception as e:
            print(f"❌ Error in datatable submission: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'❌ An error occurred: {str(e)}'
            }, status=500)

    def handle_file_upload(self, request):
        """Handle file upload (existing functionality)"""
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "❌ No file uploaded.")
            return Response({'master_data': RecoveryMasterCreation.objects.all()}, status=status.HTTP_400_BAD_REQUEST)

        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid.")
            return Response({'master_data': RecoveryMasterCreation.objects.all()}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            sheet = wb.active
            if not sheet:
                messages.error(request, "❌ Could not read the Excel sheet.")
                return Response({'master_data': RecoveryMasterCreation.objects.all()}, status=status.HTTP_400_BAD_REQUEST)

            success_count = 0
            failure_count = 0
            failed_rows = []

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                print(f"\n🔄 Processing row: {row}")

                if len(row) < 5:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ⚠️ Skipped — Less than 5 columns.")
                    continue

                stock_raw = str(row[0]).strip() if row[0] else ''
                model_field = str(row[1]).strip() if row[1] else ''
                plating_color_text = str(row[2]).strip() if row[2] else ''
                qty = row[3]
                s_loc = str(row[4]).strip() if row[4] else ''

                empty_fields = []
                if not stock_raw:
                    empty_fields.append("Stock No")
                if not model_field:
                    empty_fields.append("Model Code")
                if not plating_color_text:
                    empty_fields.append("Plating Color")
                if qty in [None, '', 0]:
                    empty_fields.append("Total Quantity")
                if not s_loc:
                    empty_fields.append("Location")

                if empty_fields:
                    field_list = ", ".join(empty_fields)
                    failed_rows.append(f"Row {idx}: ❌ {field_list} should not be empty.")
                    failure_count += 1
                    continue

                model_stock, codes, error_msg = self.validate_codes(stock_raw, model_field)
                if error_msg:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: {error_msg}")
                    continue

                if not codes:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Could not extract codes from model field.")
                    continue
                plating_code, polish_code, version_code = codes

                print(f"🧪 Model Code Extracted ➤ Plating: '{plating_code}', Polish: '{polish_code}', Version: '{version_code}'")

                plating_obj_code = Plating_Color.objects.filter(plating_color_internal=plating_code).first()
                if not plating_obj_code:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Plating code '{plating_code}' not found.")
                    continue

                plating_color_obj_excel = Plating_Color.objects.filter(plating_color=plating_color_text).first()
                if not plating_color_obj_excel:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Excel Plating Color '{plating_color_text}' not found.")
                    continue

                polish_obj = PolishFinishType.objects.filter(polish_internal=polish_code).first()
                if not polish_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Polish code '{polish_code}' not found.")
                    continue

                version_obj = Version.objects.filter(version_internal=version_code).first()
                if not version_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Version code '{version_code}' not found.")
                    continue

                if "_" not in s_loc:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Invalid s_loc format: '{s_loc}'")
                    continue

                vendor_name, loc_name = s_loc.split("_", 1)
                vendor_obj = Vendor.objects.filter(vendor_name=vendor_name).first()
                if not vendor_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Vendor '{vendor_name}' not found.")
                    continue

                location_obj = Location.objects.filter(location_name=loc_name).first()
                if not location_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Location '{loc_name}' not found.")
                    continue

                batch_id = f"BATCH-{datetime.now().strftime('%Y%m%d%H%M%S')}-{idx}"

                try:
                    RecoveryMasterCreation.objects.create(
                        batch_id=batch_id,
                        model_stock_no=model_stock,
                        plating_color=plating_color_obj_excel.plating_color,
                        vendor_internal=vendor_obj.vendor_internal,
                        location=location_obj,
                        tray_capacity=model_stock.tray_capacity if model_stock else None,
                        tray_type=model_stock.tray_type.tray_type if model_stock and model_stock.tray_type else None,
                        gender=model_stock.gender if model_stock else None,
                        ep_bath_type=model_stock.ep_bath_type if model_stock else None,
                        brand=model_stock.brand if model_stock else None,
                        total_batch_quantity=qty,
                        version=version_obj if version_obj else None,
                        polish_finish=polish_obj if polish_obj else None,
                    )
                    print("✅ Row saved successfully!")
                    success_count += 1
                except Exception as e:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Error saving row: {str(e).splitlines()[0]}")

            if success_count > 0 and failure_count == 0:
                messages.success(request, f"{success_count} row(s) uploaded successfully.")
            elif success_count > 0 and failure_count > 0:
                error_msg = f"⚠️ Partial Success: {success_count} succeeded, {failure_count} failed.\n\n"
                error_msg += "\n".join(failed_rows)
                messages.warning(request, mark_safe(error_msg.replace('\n', '<br>')))
            elif success_count == 0 and failure_count > 0:
                error_msg = f"❌ Upload Failed: All {failure_count} row(s) failed.\n\n"
                error_msg += "\n".join(failed_rows)
                messages.error(request, mark_safe(error_msg.replace('\n', '<br>')))

            return Response({'master_data': RecoveryMasterCreation.objects.all()})

        except Exception as e:
            print(f"❌ Error processing file: {str(e)}")
            messages.error(request, f"❌ An error occurred: {str(e)}")
            return Response({'master_data': RecoveryMasterCreation.objects.all()}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class REC_DPBulkUploadPreviewView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'success': False, 'error': 'No file uploaded.'}, status=400)
        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            return Response({'success': False, 'error': 'Invalid file type.'}, status=400)
        try:
            wb = openpyxl.load_workbook(uploaded_file)
            sheet = wb.active
            # Read header row
            header = [str(cell.value).strip().lower() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            required = ['input_stock_no', 'model', 'plating_color', 'qty', 's_loc']
            if header[:len(required)] != required:
                return Response({
                    'success': False,
                    'error': 'Excel file must have columns: input_stock_no, model, plating_color, qty, s_loc (in this order, first row)'
                }, status=400)
            data = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                if not any(row):
                    continue
                data.append({
                    'ID': str(idx),
                    'Stock No': row[0] or '',
                    'Model': row[1] or '',
                    'Plating Color': row[2] or '',
                    'Quantity': row[3] or '',
                    'S Loc': row[4] or '',
                })
            return Response({'success': True, 'data': data})
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=500)
        
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.permissions import IsAuthenticated
import math
import json

class REC_PickTableAPIView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Recovery_DP/Recovery_DP_PickTable.html'

    def get(self, request, *args, **kwargs):
        user = request.user

        # Subqueries for annotations
        last_process_module_subquery = RecoveryStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('last_process_module')[:1]
        next_process_module_subquery = RecoveryStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('next_process_module')[:1]

        # Build queryset
        queryset = RecoveryMasterCreation.objects.filter(
            total_batch_quantity__gt=0
        ).annotate(
            last_process_module=Subquery(last_process_module_subquery),
            next_process_module=Subquery(next_process_module_subquery),

        )

        # Pagination
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 10)  # 10 items per page
        page_obj = paginator.get_page(page_number)

        # Convert page_obj to list of dicts (using .values())
        master_data = list(page_obj.object_list.values(
            'batch_id',
            'date_time',
            'model_stock_no__model_no',
            'plating_color',
            'polish_finish',
            'version__version_name',
            'vendor_internal',
            'location__location_name',
            'no_of_trays',
            'tray_type',
            'total_batch_quantity',
            'tray_capacity',
            'Moved_to_R_Picker',
            'last_process_module',
            'next_process_module',
            'R_Draft_Saved',
            'rec_pick_remarks'
        ))

        # Calculate no_of_trays dynamically
        for data in master_data:
            total_batch_quantity = data.get('total_batch_quantity', 0)
            tray_capacity = data.get('tray_capacity', 0)
            data['vendor_location'] = f"{data.get('vendor_internal', '')}_{data.get('location__location_name', '')}"
            if tray_capacity > 0:
                no_of_trays = math.ceil(total_batch_quantity / tray_capacity)
                data['no_of_trays'] = no_of_trays
                tray_qty_list = []
                remainder = total_batch_quantity % tray_capacity
                if no_of_trays == 1:
                    tray_qty_list = [total_batch_quantity]
                elif no_of_trays > 1:
                    if remainder != 0:
                        tray_qty_list.append(remainder)
                        for _ in range(1, no_of_trays):
                            tray_qty_list.append(tray_capacity)
                    else:
                        for _ in range(no_of_trays):
                            tray_qty_list.append(tray_capacity)
                data['tray_qty_list'] = tray_qty_list
                
                
            else:
                data['no_of_trays'] = 0
                data['tray_qty_list'] = []
                
           # Add model images
            mmc = RecoveryMasterCreation.objects.filter(batch_id=data['batch_id']).first()
            images = []
            if mmc:
                model_master = mmc.model_stock_no
                for img in getattr(model_master, 'images', []).all():
                    if getattr(img, 'master_image', None):
                        images.append(img.master_image.url)
            if not images:
                from django.templatetags.static import static
                images = [static('assets/images/imagePlaceholder.png')]
            data['model_images'] = images

  
        context = {
            'master_data': master_data,
            'page_obj': page_obj,
            'paginator': paginator,
            'user': user,
        }
        return Response(context, template_name=self.template_name)

@method_decorator(csrf_exempt, name='dispatch')
class Rec_TrayIdScanAPIView(APIView):
    def post(self, request):
        try:
            # Parse JSON data
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            trays = data.get('trays', [])
            user = request.user if request.user.is_authenticated else None

            # You may want to send these from the frontend or fetch from batch
            lot_id = data.get('lot_id') or self.generate_new_lot_id()
            # Optionally, get other fields from batch/model if not sent

            if not batch_id or not trays:
                return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)

            batch_instance = RecoveryMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({'success': False, 'error': 'Invalid batch_id.'}, status=400)

            # Fetch related fields from batch_instance or data
            model_stock_no = batch_instance.model_stock_no
            version = batch_instance.version
            total_batch_quantity = batch_instance.total_batch_quantity
            polish_finish_obj = PolishFinishType.objects.filter(polish_finish=batch_instance.polish_finish).first() 
            plating_color = batch_instance.plating_color

            # Save both models in a transaction
            with transaction.atomic():
                # Save RecoveryStockModel
                total_stock_obj = RecoveryStockModel.objects.create(
                    batch_id=batch_instance,
                    model_stock_no=model_stock_no,
                    version=version,
                    total_stock=total_batch_quantity,
                    polish_finish=polish_finish_obj,
                    plating_color=Plating_Color.objects.filter(plating_color=plating_color).first() if plating_color else None,
                    lot_id=lot_id,
                    tray_scan_status=True,
                    last_process_date_time=now(),
                    last_process_module="DayPlanning",
                    next_process_module="IP Screening",
                )

                # Save all TrayId rows
                for tray in trays:
                    tray_id = tray.get('tray_id')
                    tray_quantity = tray.get('tray_quantity')
                    
                    if not tray_id or not tray_quantity:
                        continue
                    RecoveryTrayId.objects.create(
                        lot_id=lot_id,
                        tray_id=tray_id,
                        batch_id=batch_instance,
                        tray_quantity=int(tray_quantity),
                        user=user
                    )

                # Update Moved_to_R_Picker
                RecoveryMasterCreation.objects.filter(batch_id=batch_id).update(Moved_to_R_Picker=True)

            return JsonResponse({'success': True, 'message': 'Tray scan and stock saved!'}, status=201)

        except IntegrityError as e:
            return JsonResponse({'success': False, 'error': 'Integrity error: ' + str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Unexpected error: ' + str(e)}, status=500)

    def generate_new_lot_id(self):
        from datetime import datetime
        timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
        last_lot = RecoveryStockModel.objects.order_by('-id').first()
        if last_lot and last_lot.lot_id and last_lot.lot_id.startswith("LID"):
            last_seq_no = int(last_lot.lot_id[-4:])
            next_seq_no = last_seq_no + 1
        else:
            next_seq_no = 1
        seq_no = f"{next_seq_no:04d}"
        return f"LID{timestamp}{seq_no}"
    

@method_decorator(csrf_exempt, name='dispatch')
class Rec_TrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        tray_ids = RecoveryTrayId.objects.filter(batch_id__batch_id=batch_id)
        data = [
            {
                's_no': idx + 1,
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity
            }
            for idx, tray in enumerate(tray_ids)
        ]
        return JsonResponse({'success': True, 'trays': data})
    
@method_decorator(csrf_exempt, name='dispatch')
class Rec_DraftTrayIdAPIView(APIView):
    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            trays = data.get('trays', [])
            user = request.user if request.user.is_authenticated else None

            if not batch_id or not trays:
                return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)

            batch_instance = RecoveryMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({'success': False, 'error': 'Invalid batch_id.'}, status=400)

            lot_id = data.get('lot_id') or f"DRAFT-{batch_id}"

            for tray in trays:
                tray_id = tray.get('tray_id')
                tray_quantity = tray.get('tray_quantity')
                draft_id = tray.get('id', None)
                if not tray_id or not tray_quantity:
                    continue
                if draft_id:
                    # Update by id if present
                    Rec_DraftTrayId.objects.filter(id=draft_id).update(
                        tray_id=tray_id,
                        tray_quantity=int(tray_quantity),
                        lot_id=lot_id,
                        batch_id=batch_instance,
                        user=user
                    )
                else:
                    # Create new if not present
                    Rec_DraftTrayId.objects.update_or_create(
                        tray_id=tray_id,
                        defaults={
                            'lot_id': lot_id,
                            'batch_id': batch_instance,
                            'tray_quantity': int(tray_quantity),
                            'user': user
                        }
                    )

            batch_instance.R_Draft_Saved = True
            batch_instance.save(update_fields=['R_Draft_Saved'])

            return JsonResponse({'success': True, 'message': 'Draft saved!'}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Unexpected error: ' + str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class Rec_DraftTrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        ids = request.GET.getlist('ids[]')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        draft_trays = Rec_DraftTrayId.objects.filter(batch_id__batch_id=batch_id)
        if ids:
            # Preserve order as per ids
            id_map = {str(tray.id): tray for tray in draft_trays}
            ordered_trays = [id_map[i] for i in ids if i in id_map]
        else:
            ordered_trays = list(draft_trays)
        data = [
            {
                'id': tray.id,
                's_no': idx + 1,
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity
            }
            for idx, tray in enumerate(ordered_trays)
        ]
        return JsonResponse({'success': True, 'trays': data})
    
@method_decorator(csrf_exempt, name='dispatch')
class Rec_TrayIdUniqueCheckAPIView(APIView):
    def get(self, request):
        tray_id = request.GET.get('tray_id')
        if not tray_id:
            return JsonResponse({'exists': False})
        exists = RecoveryTrayId.objects.filter(tray_id=tray_id).exists()
        return JsonResponse({'exists': exists})

@method_decorator(csrf_exempt, name='dispatch')
class Rec_UpdateBatchQuantityAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            new_quantity = data.get('total_batch_quantity')
            if not batch_id or new_quantity is None:
                return JsonResponse({'success': False, 'error': 'Missing batch_id or quantity'}, status=400)
            obj = RecoveryMasterCreation.objects.filter(batch_id=batch_id, Moved_to_R_Picker=False).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found or already moved'}, status=404)
            obj.total_batch_quantity = new_quantity
            obj.save(update_fields=['total_batch_quantity'])
            return JsonResponse({'success': True, 'message': 'Quantity updated'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class Rec_DeleteBatchAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            obj = RecoveryMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)
            obj.delete()
            return JsonResponse({'success': True, 'message': 'Batch deleted'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class Rec_SaveDPPickRemarkAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            remark = data.get('remark', '').strip()
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            obj = RecoveryMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)
            obj.rec_pick_remarks = remark
            obj.save(update_fields=['rec_pick_remarks'])
            return JsonResponse({'success': True, 'message': 'Remark saved'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class Rec_DPCompletedTableView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Day_Planning/DP_Completed_Table.html'

    def get(self, request, *args, **kwargs):
        user = request.user

        # Subqueries for annotations
        last_process_module_subquery = RecoveryStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('last_process_module')[:1]
        next_process_module_subquery = RecoveryStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('next_process_module')[:1]

        # Build queryset
        queryset = RecoveryMasterCreation.objects.filter(
            total_batch_quantity__gt=0,
            Moved_to_R_Picker=True  # Only include records where Moved_to_R_Picker is True
        ).annotate(
            last_process_module=Subquery(last_process_module_subquery),
            next_process_module=Subquery(next_process_module_subquery),
        )

        # Pagination
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 10)  # 10 items per page
        page_obj = paginator.get_page(page_number)

        # Convert page_obj to list of dicts (using .values())
        master_data = list(page_obj.object_list.values(
            'batch_id',
            'date_time',
            'model_stock_no__model_no',
            'plating_color',
            'polish_finish',
            'version__version_name',
            'vendor_internal',
            'location__location_name',
            'no_of_trays',
            'tray_type',
            'total_batch_quantity',
            'tray_capacity',
            'Moved_to_R_Picker',
            'last_process_module',
            'next_process_module',
            'R_Draft_Saved',
        ))

        # Calculate no_of_trays dynamically
        for data in master_data:
            total_batch_quantity = data.get('total_batch_quantity', 0)
            tray_capacity = data.get('tray_capacity', 0)
            data['vendor_location'] = f"{data.get('vendor_internal', '')}_{data.get('location__location_name', '')}"
            if tray_capacity > 0:
                data['no_of_trays'] = math.ceil(total_batch_quantity / tray_capacity)
            else:
                data['no_of_trays'] = 0
                
            # Add model images
            mmc = RecoveryMasterCreation.objects.filter(batch_id=data['batch_id']).first()
            images = []
            if mmc:
                model_master = mmc.model_stock_no
                for img in getattr(model_master, 'images', []).all():
                    if getattr(img, 'master_image', None):
                        images.append(img.master_image.url)
            if not images:
                from django.templatetags.static import static
                images = [static('assets/images/imagePlaceholder.png')]
            data['model_images'] = images

        context = {
            'master_data': master_data,
            'page_obj': page_obj,
            'paginator': paginator,
            'user': user,
        }
        return Response(context, template_name=self.template_name)
    